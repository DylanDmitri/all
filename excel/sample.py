from openpyxl import Workbook
import datetime

wb = Workbook()


ws = wb.active

ws['A1'] = 42


ws['A2'] = datetime.datetime.now()

ws.append([1, 2, 3])



wb.save("one_button_one_light.xlsx")
